#!/usr/bin/env python
# Please update inventory.csv for IP and Credentials
import csv
import sys
import paramiko
import time
import re
import ipaddress
import json
from openpyxl import load_workbook
from openpyxl import Workbook


def excel_to_lists(filename):
    wb = load_workbook(filename)
    ws = wb.active
    invents = list()
    for i in range(2, ws.max_row+1):
        invent_dict = {
            'hostname': None,
            'address': None,
            'type' : None,
            'username': None,
            'password': None
            }

        invent_dict['hostname'] = ws.cell(row=i, column=1).value
        invent_dict['address'] = ws.cell(row=i, column=2).value
        invent_dict['type'] = ws.cell(row=i, column=3).value
        invent_dict['username'] = ws.cell(row=i, column=4).value
        invent_dict['password'] = ws.cell(row=i, column=5).value
        invents.append(invent_dict)
    return invents
   
success = 0
failure = 0
iteration = 0
inventory_file = input("Please input the path of XLSX file (Ex: D:\inventory.xlsx): ")
print('-'*80)
device_list = excel_to_lists(inventory_file)
#print(json.dumps(output, indent=4))

ssh_client = paramiko.SSHClient()
ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
for device in device_list:
    iteration += 1
    try:
        ipaddress.ip_address(device["address"])
    except ValueError:
        print("Invalid IP Address Format : {} --> Please check IP Address on row number {}!".format(device["address"],iteration))
        failure += 1
        continue
    device_info = { "ip": device["address"],
                    "username":device["username"],
                    "password":device["password"],
                  }
    #print(device_info)    
    try:
        ssh_client.connect(device_info["ip"],username=device_info["username"],password=device_info["password"])
    except:
        print("SSH Connection Error for {}. Please check your connection or credentials!".format(device_info["ip"]))
        failure += 1
        continue
    print("Successfully connect to {}".format(device_info["ip"]))
    success += 1
    ssh_conn = ssh_client.invoke_shell()
    ssh_conn.send("terminal length 0\n")
    ssh_conn.send("show run")
    ssh_conn.send("\n")
    time.sleep(2)
    if ssh_conn.recv_ready():
        output = ssh_conn.recv(65535)
    # Python3 treat default to binary raw bits, not string implicitly --> use .decode()
    output_str=str(output.decode())
    output_str = re.search(r"!.*end",output_str,flags=re.DOTALL).group(0)
    with open("{}.cfg".format(device_info["ip"]), mode="w") as f:
        f.write(output_str)
    print("Successfully backup device {}".format(device_info["ip"]))
    print('-'*80) 
f.close()
print("Attempt to backup {} devices".format(success+failure))
print("Success : {}".format(success))
print("Failure : {}".format(failure))
print('-'*80)


